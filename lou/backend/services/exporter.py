import json
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.core.schema import Rule, Commit, ChangeType

RULE_TYPE_FILLS = {
    "standard":   PatternFill("solid", fgColor="C6EFCE"),   # Standard rows get green.
    "fallback":   PatternFill("solid", fgColor="FFEB9C"),   # Fallback rows get amber.
    "red_line":   PatternFill("solid", fgColor="FFC7CE"),   # Red lines need to stand out.
    "escalation": PatternFill("solid", fgColor="BDD7EE"),   # Escalation rows get blue.
}

CHANGE_TYPE_FILLS = {
    "confirms":    PatternFill("solid", fgColor="E2EFDA"),  # A confirmed rule gets soft green.
    "contradicts": PatternFill("solid", fgColor="FFDDD8"),  # A conflict gets soft red.
    "new_rule":    PatternFill("solid", fgColor="D9E1F2"),  # New-rule proposals get soft blue.
    "manual":      PatternFill("solid", fgColor="EDEDED"),  # Manual edits stay neutral.
    "initial":     PatternFill("solid", fgColor="FFFFFF"),
    "extends":     PatternFill("solid", fgColor="FFF2CC"),  # Extensions get soft yellow.
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max(12, max_len + 2)


def generate_excel(rules: list[Rule], commits: list[Commit]) -> bytes:
    wb = Workbook()

    # Sheet 1 is the current playbook.
    ws_rules = wb.active
    ws_rules.title = "Rules"
    rule_headers = [
        "Rule ID", "Topic", "Category", "Rule Type",
        "Standard Position", "Fallback Position", "Red Line",
        "Reasoning", "Suggested Language", "Decision Logic",
        "Sources", "Confidence", "Version", "Committed By", "Last Updated",
    ]
    for col, h in enumerate(rule_headers, 1):
        ws_rules.cell(row=1, column=col, value=h)
        _style_header(ws_rules, 1, col)

    for row_idx, rule in enumerate(rules, 2):
        sources = json.loads(rule.sources) if isinstance(rule.sources, str) else rule.sources
        values = [
            rule.rule_id, rule.topic, rule.category, rule.rule_type.value if hasattr(rule.rule_type, "value") else rule.rule_type,
            rule.standard_position, rule.fallback_position or "",
            rule.red_line or "", rule.reasoning,
            rule.suggested_language or "", rule.decision_logic or "",
            ", ".join(sources), f"{rule.confidence:.0%}",
            rule.version, rule.committed_by,
            rule.committed_at.strftime("%Y-%m-%d %H:%M") if rule.committed_at else "",
        ]
        rt = rule.rule_type.value if hasattr(rule.rule_type, "value") else str(rule.rule_type)
        fill = RULE_TYPE_FILLS.get(rt)
        for col_idx, val in enumerate(values, 1):
            cell = ws_rules.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN
            if col_idx == 4 and fill:  # Color the Rule Type cell.
                cell.fill = fill

    ws_rules.freeze_panes = "A2"
    _autofit(ws_rules)

    # Sheet 2 is the commit log.
    ws_log = wb.create_sheet("Changelog")
    log_headers = [
        "Timestamp", "Rule ID", "Change Type", "Changed By",
        "Source Document", "Lawyer Note", "Old Position", "New Position",
    ]
    for col, h in enumerate(log_headers, 1):
        ws_log.cell(row=1, column=col, value=h)
        _style_header(ws_log, 1, col)

    for row_idx, commit in enumerate(sorted(commits, key=lambda c: c.committed_at, reverse=True), 2):
        old_val = ""
        new_val = ""
        try:
            old_data = json.loads(commit.old_value) if commit.old_value else {}
            old_val = old_data.get("standard_position", "")[:300]
        except Exception:
            pass
        try:
            new_data = json.loads(commit.new_value) if commit.new_value else {}
            new_val = new_data.get("standard_position", "")[:300]
        except Exception:
            pass

        ct = commit.change_type.value if hasattr(commit.change_type, "value") else str(commit.change_type)
        values = [
            commit.committed_at.strftime("%Y-%m-%d %H:%M") if commit.committed_at else "",
            commit.rule_id, ct, commit.committed_by,
            commit.source_document or "", commit.lawyer_note or "",
            old_val, new_val,
        ]
        fill = CHANGE_TYPE_FILLS.get(ct)
        for col_idx, val in enumerate(values, 1):
            cell = ws_log.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN
            if col_idx == 3 and fill:
                cell.fill = fill

    ws_log.freeze_panes = "A2"
    _autofit(ws_log)

    # Sheet 3 shows which documents fed the playbook.
    ws_src = wb.create_sheet("Sources")
    src_headers = ["Source Document", "Rules Informed", "Upload Date", "Uploaded By"]
    for col, h in enumerate(src_headers, 1):
        ws_src.cell(row=1, column=col, value=h)
        _style_header(ws_src, 1, col)

    # Build one row per source document.
    source_map: dict[str, dict] = {}
    for commit in commits:
        doc = commit.source_document or "Playbook"
        if doc not in source_map:
            source_map[doc] = {
                "rules": set(),
                "date": commit.committed_at,
                "by": commit.committed_by,
            }
        source_map[doc]["rules"].add(commit.rule_id)
        if commit.committed_at and commit.committed_at < source_map[doc]["date"]:
            source_map[doc]["date"] = commit.committed_at

    for row_idx, (doc, info) in enumerate(source_map.items(), 2):
        values = [
            doc,
            ", ".join(sorted(info["rules"])),
            info["date"].strftime("%Y-%m-%d") if info["date"] else "",
            info["by"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_src.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN

    ws_src.freeze_panes = "A2"
    _autofit(ws_src)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
